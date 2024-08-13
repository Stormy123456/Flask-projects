from flask import Flask, render_template, request, redirect, Response, jsonify, send_file, session
from datetime import datetime, timedelta, time
from config_db import connection_string, connection_string_invoice
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Border, Side
import pyodbc 
import pandas as pd
import ast
import sys
import codecs

def api_get_datapicking():
    data = []
    excel_file = []
    data_excel = []
    data_excel_to_pgsql = []
    data_unique_from_excel = []
    data_unique_from_query = []
    data_qty_unique = []
    data_sku_unique = []
    data_not_equal_from_excel = []
    data_not_equal_from_data = []
    stamp_invoice_both_have = []
    invoice_excel_line_unequal = [] 
    max_date = None
    min_date = None
    
    file = request.files.get('file')  # อ่านไฟล์ที่ถูกอัปโหลด
    df = pd.read_excel(file, na_values=[' '])
    df = df.where(pd.notna(df), None)
    excel_file = df.to_dict(orient='records')
    # สร้าง defaultdict เพื่อเก็บจำนวน lines สำหรับแต่ละ invoice_number
    lines_count_dict = defaultdict(int)
    for record in excel_file:
        po_invoice_value = record.get('PO / Invoice')
        cleaned_value = po_invoice_value.split(' / ')[0].strip()
        date = record.get('Date')
        sku = record.get('SKU')
        material_description = record.get('Material description')
        qty = record.get('Qty In EWM')
        weight = record.get('Weight')
        suppliername = record.get('SupplierName')
        doctype = record.get('DocType')
        sloc = record.get('Sloc')
        serial = record.get('Serial')
        lines_count_dict[cleaned_value] += 1
                    
        date_check = datetime.strptime(date, '%d.%m.%Y %H:%M:%S')
        if max_date is None or date_check > max_date:
            max_date = date_check
                        
        if min_date is None or date_check < min_date:
            min_date = date_check

        data_excel.append({
            'invoice_number': str(cleaned_value),
            'create_on': str(date),
            'barcode': str(sku),
            'product': str(material_description),
            'qty': str(qty),
            'weight': str(weight),
            'contact':str(suppliername),
            'order_type_name': str(doctype),
            'internal_reference': str(sloc),
            'serial_number': str(serial),
            'count': lines_count_dict[cleaned_value]  # เพิ่มคอลัมน์ lines_count
        })
        
    if max_date.time() > time(6, 0):
        max_date = max_date.strftime('%Y-%m-%d 23:59:59')
    min_date = min_date.strftime('%Y-%m-%d 00:00:00')
    if min_date and max_date:
        if file.filename == '':
            return render_template('index_stock_picking_automate.html', data=data, error_message="ไฟลไม่มีข้อมูล")
        else:
            # SQL 
            try:
                connection = pyodbc.connect(connection_string)
                cursor = connection.cursor()
                ############################################################ So query ############################################################
                sql_query = f"""SELECT 
                                sp.name AS reference,
                                TO_CHAR(sp.create_date + interval '7 hours', 'DD.MM.YYYY HH24:MI:SS') AS created_on,
                                pp.default_code AS barcode,
                                pp.product_name AS product,
                                pp.weight AS weight,
                                ROUND(sm.product_uom_qty) AS qty,
                                rp.name AS contact,
                                spt.name AS order_type_name,
                                sl.name AS source_location,
                                sl.internal_reference AS internal_reference,
                                sml.imei AS serial_number,
                                sp.api_state as api_status,
                                CASE
                                    WHEN sp.state = 'draft' THEN 'Draft'
                                    WHEN sp.state = 'waiting' THEN 'Waiting Another Operation'
                                    WHEN sp.state = 'confirmed' THEN 'Waiting'
                                    WHEN sp.state = 'assigned' THEN 'Ready'
                                    WHEN sp.state = 'done' THEN 'Done'
                                    WHEN sp.state = 'cancel' THEN 'Cancelled'
                                END AS status,
                                COUNT(*) OVER (PARTITION BY sp.name) AS total_lines
                            FROM 
                                stock_picking sp 
                            LEFT JOIN 
                                sale_purchase_types spt ON spt.id = sp.sale_purchase_type_id 
                            LEFT JOIN 
                                stock_location sl ON sl.id = sp.location_id 
                            LEFT JOIN 
                                stock_location sld ON sld.id = sp.location_dest_id 
                            LEFT JOIN 
                                res_partner rp ON rp.id = sp.partner_id 
                            LEFT JOIN 
                                stock_picking_batch spb ON spb.id = sp.batch_id 
                            LEFT JOIN 
                                stock_move sm ON sm.picking_id = sp.id
                            LEFT JOIN 
                                ( 
                                    SELECT 
                                        sml_sub.id,
                                        sml_sub.lot_id,
                                        sml_sub.move_id,
                                        sml_sub.qty_done,
                                        s_pro_l.name as imei,
                                        ROW_NUMBER() OVER (PARTITION BY sml_sub.id ORDER BY sml_sub.id) AS rn
                                    FROM
                                        stock_move_line sml_sub
                                    LEFT JOIN 
                                		stock_production_lot s_pro_l ON s_pro_l.id = sml_sub.lot_id  
                                ) AS sml ON sml.move_id = sm.id AND rn = 1 AND sml.imei IS NOT NULL                             
                            LEFT JOIN 
                                product_product pp ON pp.id = sm.product_id 
                            WHERE 
                                (sp.api_state = 'send' OR sp.api_state = 'fail')
                                AND sp.name LIKE '%OUT%'
                                AND sp.create_date + interval '7 hours' >= '{min_date}'
                                AND sp.create_date + interval '7 hours' <= '{max_date}'
                    """
                cursor.execute(sql_query)
                result = cursor.fetchall()
                data = [{
                            'invoice_number': str(item.reference),
                            'create_on': str(item.created_on),
                            'barcode': str(item.barcode),
                            'product': str(item.product),
                            'qty': str(item.qty),
                            'weight': str(item.weight),
                            'contact': str(item.contact),
                            'order_type_name': str(item.order_type_name),
                            'internal_reference': str(item.internal_reference),
                            'serial_number': str(item.serial_number),
                            'count': item.total_lines,
                            'status_tg': item.status
                        } for item in result]
                print("Picking is Success")
            except Exception as e:
                print("Picking Error:", str(e))
            finally:
                connection.close()  # ปิด Connection

        data_result = pd.DataFrame(data)
        data_result['qty'] = pd.to_numeric(data_result['qty'])
        result = data_result.groupby('invoice_number')['qty'].sum().reset_index()
        
        excel_data_result = pd.DataFrame(data_excel)
        excel_data_result['qty'] = pd.to_numeric(excel_data_result['qty'])
        excel_result = excel_data_result.groupby(['invoice_number', 'barcode'])['qty'].sum().reset_index()
        result_list_from_excel = excel_result.to_dict(orient='records')
        
        data_array = {(item['invoice_number']) for item in data}
        data_set_array = list(data_array)
        # สร้างเซตของค่า key1 และ key2 จาก array2
        excel_array = {(item['invoice_number']) for item in data_excel}
        excel_set_array = list(excel_array)
        
        check_first_time_invoice = True
        check_first_time_barcode = True
        first_time_invoice = ''
        first_time_barcode = ''
        for excel_item in data_excel:
            if excel_item['invoice_number'] == first_time_invoice and excel_item['barcode'] == first_time_barcode:
                check_first_time_invoice = False
                check_first_time_barcode = False
            else:
                first_time_invoice = excel_item['invoice_number']
                check_first_time_invoice = True
                first_time_barcode = excel_item['barcode']
                check_first_time_barcode = True
            if excel_item['invoice_number'] not in data_set_array:
                excel_item['differences'] = 'ไม่พบบิลนี้ใน Odoo'
                data_unique_from_excel.append({'invoice_number': excel_item['invoice_number'],'create_on': excel_item['create_on']})
                data_excel_to_pgsql.append({'type_name': 'SCG','invoice_number': excel_item['invoice_number'],'create_on': excel_item['create_on'],'barcode': excel_item['barcode'],'product': excel_item['product'],'qty': excel_item['qty'],'weight': excel_item['weight'],'contact': excel_item['contact'],'order_type_name': excel_item['order_type_name'],'internal_reference': excel_item['internal_reference'],'serial_number': excel_item['serial_number'],'count': excel_item['count'],'status_tg': ''})
            else:
                qty_excel = 0
                for excel_qty in result_list_from_excel:
                    if excel_qty['invoice_number'] == excel_item['invoice_number'] and excel_qty['barcode'] == excel_item['barcode']:
                        qty_excel = excel_qty['qty']

                is_barcode_unique = True
                is_contact_unique = True
                is_internal_reference = True
                is_serial_number_unique = True
                data_qty_unique_check_sku = []
                for check_data in data:
                    if check_data['invoice_number'] == excel_item['invoice_number']:
                        if excel_item['barcode'] == check_data['barcode'] and check_data['barcode'] is not None:
                            is_barcode_unique = False
                        if excel_item['contact'] == check_data['contact'] and check_data['contact'] is not None:
                            is_contact_unique = False
                        if excel_item['internal_reference'] == check_data['internal_reference'] and check_data['internal_reference'] is not None:
                            is_internal_reference = False
                        if excel_item['serial_number'] == check_data['serial_number'] and check_data['serial_number'] is not None:
                            is_serial_number_unique = False  
                        if check_data['barcode'] != excel_item['barcode'] and check_data['serial_number'] == excel_item['serial_number'] and check_data['serial_number'] != 'None' and excel_item['serial_number'] != 'None':
                            data_sku_unique.append({
                                'product_tg': check_data['product'],
                                'barcode_tg': check_data['barcode'],
                                'serial_number_tg': check_data['serial_number'],
                                'product_scg': excel_item['product'],
                                'barcode_scg': excel_item['barcode'],
                                'serial_number_scg': excel_item['serial_number']
                            })
                        if excel_item['barcode'] == check_data['barcode'] and  int(qty_excel) != int(check_data['qty']) and check_first_time_invoice == True and check_first_time_barcode == True:
                            found = False
                            # หากไม่พบ barcode เดียวกันใน list ให้เพิ่มข้อมูลเข้า list
                            for item in data_qty_unique_check_sku:
                                if item['invoice_number'] == excel_item['invoice_number'] and item['barcode'] == excel_item['barcode']:
                                    # หากพบ barcode เหมือนกันใน list ให้เพิ่ม qty เข้าไป
                                    qty_tg = int(item['qty_tg'])
                                    qty_tg += int(check_data['qty'])
                                    item['qty_tg'] = qty_tg
                                    found = True
                                    break
                            if found == False:
                                data_qty_unique_check_sku.append({
                                    'invoice_number': excel_item['invoice_number'],
                                    'barcode': excel_item['barcode'],
                                    'product': excel_item['product'],
                                    'qty_tg': int(check_data['qty']),
                                    'qty_scg': qty_excel,
                                    'weight': excel_item['weight'],
                                    'contact': excel_item['contact'],
                                    'order_type_name': excel_item['order_type_name'],
                                    'internal_reference': excel_item['internal_reference']
                                })
                                
                            
                            
                # ตรวจสอบว่ามีข้อมูลใน data ที่มี invoice เหมือนกัน barcode และ qty เหมือนกันหรือไม่
                for check_data_sku in data_qty_unique_check_sku[:]:
                    if check_data_sku['invoice_number'] == excel_item['invoice_number'] and check_data_sku['barcode'] == excel_item['barcode']:
                        # ถ้า qty_tg เท่ากับ qty_scg ให้ลบข้อมูลนี้ออก
                        if check_data_sku['qty_tg'] == int(qty_excel):  # แก้เปรียบเทียบให้เปรียบเทียบ qty_tg กับ qty_excel เป็นจำนวนเต็ม
                            data_qty_unique_check_sku.remove(check_data_sku)

                data_qty_unique.extend(data_qty_unique_check_sku)
                if is_barcode_unique == True or is_contact_unique == True or is_internal_reference == True or is_serial_number_unique == True:
                    data_not_equal_from_excel.append(excel_item)


        # สร้าง DataFrame จาก data_excel
        df = pd.DataFrame(data_excel)

        # Group by invoice_number และนับจำนวน invoice_number ที่ซ้ำกัน
        invoice_counts = df.groupby('invoice_number').size().reset_index(name='count')
        for index, row in invoice_counts.iterrows():
            first_invoice = ''
            current_invoice = ''
            for data_sql_check in data:
                first_invoice = data_sql_check['invoice_number']
                if first_invoice != current_invoice:
                    current_invoice = first_invoice
                    if str(row['invoice_number']) == str(data_sql_check['invoice_number']) and int(row['count']) != int(data_sql_check['count']):
                        stamp_invoice_both_have.append({
                            'invoice_number': row['invoice_number'],
                            'tg_lines': data_sql_check['count'],
                            'scg_lines': row['count']
                        })
        # Find excel rows that invoice is same but lines not equal
        for inv in stamp_invoice_both_have:
            for excel_item in data_excel:
                if inv['invoice_number'] == excel_item['invoice_number']:
                    invoice_excel_line_unequal.append({'type_name': 'SCG','invoice_number': excel_item['invoice_number'],'create_on': excel_item['create_on'],'barcode': excel_item['barcode'],'product': excel_item['product'],'qty': excel_item['qty'],'weight': excel_item['weight'],'contact': excel_item['contact'],'order_type_name': excel_item['order_type_name'],'internal_reference': excel_item['internal_reference'],'serial_number': excel_item['serial_number'],'status_tg': ''})
        # Find unique items in data
        for data_item in data:
            if data_item['invoice_number'] not in excel_set_array:
                data_item['differences'] = 'ไม่พบบิลนี้ใน SCG'
                data_unique_from_query.append({'invoice_number': data_item['invoice_number'],'create_on': data_item['create_on']})  
                data_excel_to_pgsql.append({'type_name': 'TG','invoice_number': data_item['invoice_number'],'create_on': data_item['create_on'],'barcode': data_item['barcode'],'product': data_item['product'],'qty': data_item['qty'],'weight': data_item['weight'],'contact': data_item['contact'],'order_type_name': data_item['order_type_name'],'internal_reference': data_item['internal_reference'],'serial_number': data_item['serial_number'],'count': data_item['count'],'status_tg': data_item['status_tg']})
            else:
                is_barcode_unique = True
                is_contact_unique = True
                is_internal_reference = True
                is_serial_number_unique = True
                for check_data in data_excel:
                    if check_data['invoice_number'] in data_item['invoice_number']:
                        if data_item['barcode'] == check_data['barcode'] and check_data['barcode'] is not None:
                            is_barcode_unique = False
                        if data_item['contact'] == check_data['contact'] and check_data['contact'] is not None:
                            is_contact_unique = False
                        if data_item['internal_reference'] == check_data['internal_reference'] and check_data['internal_reference'] is not None:
                            is_internal_reference = False
                        if data_item['serial_number'] == check_data['serial_number'] and check_data['serial_number'] is not None:
                            is_serial_number_unique = False  
                if is_barcode_unique == True or is_contact_unique == True or is_internal_reference == True or is_serial_number_unique == True:
                    data_not_equal_from_data.append(data_item) 
           
        data_unique_from_excel = [dict(item) for item in set(tuple(item.items()) for item in data_unique_from_excel)]
        data_unique_from_query = [dict(item) for item in set(tuple(item.items()) for item in data_unique_from_query)]
        
        try:
            connection_invoice = pyodbc.connect(connection_string_invoice)
            cursor_inv = connection_invoice.cursor()
            emp_code = session['employee_code']
            sql_delete_query = "DELETE FROM invoice_excel WHERE emp_code = ?"
            cursor_inv.execute(sql_delete_query, (emp_code,))
            for item in data_excel_to_pgsql:
                emp_code = session['employee_code']
                type_name = item['type_name']
                invoice_number = item['invoice_number']
                create_on = item['create_on']
                barcode = item['barcode']
                product = item['product']
                qty = item['qty']
                weight = item['weight']
                contact = item['contact']
                order_type_name = item['order_type_name']
                internal_reference = item['internal_reference']
                serial_number = item['serial_number']
                count = item['count']
                status_tg = item['status_tg']
                # สร้างคำสั่ง SQL สำหรับแทรกข้อมูลลงในฐานข้อมูล
                sql_query = f"""
                    INSERT INTO invoice_excel (emp_code, type_name, invoice_number, create_on, barcode, product, qty, weight, contact, order_type_name, internal_reference, serial_number, count, status_tg)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                # Execute SQL query
                cursor_inv.execute(sql_query, (emp_code, type_name, invoice_number, create_on, barcode, product, qty, weight, contact, order_type_name, internal_reference, serial_number, count, status_tg))
            # Commit เปลี่ยนแปลง
            connection_invoice.commit()
        except Exception as e:
            print("Error:", str(e))
        finally:
            connection_invoice.close()   
            
        try:
            connection_invoice_line_not_equal = pyodbc.connect(connection_string_invoice)
            cursor_inv_not_equal = connection_invoice_line_not_equal.cursor()
            emp_code = session['employee_code']
            sql_delete_query = "DELETE FROM invoice_excel_line_not_equal WHERE emp_code = ?"
            cursor_inv_not_equal.execute(sql_delete_query, (emp_code,))
            for item in invoice_excel_line_unequal:
                emp_code = session['employee_code']
                invoice_number = item['invoice_number']
                create_on = item['create_on']
                barcode = item['barcode']
                product = item['product']
                qty = item['qty']
                weight = item['weight']
                contact = item['contact']
                order_type_name = item['order_type_name']
                internal_reference = item['internal_reference']
                serial_number = item['serial_number']
                status_tg = item['status_tg']
                sql_query = f"""
                    INSERT INTO invoice_excel_line_not_equal (emp_code, invoice_number, create_on, barcode, product, qty, weight, contact, order_type_name, internal_reference, serial_number, status_tg)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                # Execute SQL query
                cursor_inv_not_equal.execute(sql_query, (emp_code, invoice_number, create_on, barcode, product, qty, weight, contact, order_type_name, internal_reference, serial_number, status_tg))
            # Commit เปลี่ยนแปลง
            connection_invoice_line_not_equal.commit()
        except Exception as e:
            print("Error:", str(e))
        finally:
            connection_invoice_line_not_equal.close()   
    
    inv_excel = []
    inv_minus_one_found = []
    for exl_data in data_unique_from_excel:
        inv_excel.append({
            exl_data['invoice_number']
        })
    inv_strings = [str(inv).strip('{}') for inv in inv_excel]

    # สร้าง string ที่แยกด้วย comma และใส่วงเล็บเหลี่ยม
    inv_sql_string = "(" + ", ".join(inv_strings) + ")"
    
    min_minus_one = min_date = date_check - timedelta(days=1)
    if inv_excel:
        try:
            connection_minus = pyodbc.connect(connection_string)
            cursor = connection_minus.cursor()
            ############################################################ So query ############################################################
            sql_query = f"""SELECT 
                            sp.name AS reference
                        FROM 
                            stock_picking sp 
                        WHERE 
                            sp.create_date + interval '7 hours' >= '{min_minus_one}'
                            and sp.create_date + interval '7 hours' <= '{max_date}'
                            and sp.name IN {inv_sql_string}
                """
            cursor.execute(sql_query)
            result = cursor.fetchall()
            inv_minus_one_found = [{
                        'invoice_number': str(item.reference)
                    } for item in result]
            print("Picking is Success")
        except Exception as e:
            print("Picking Error:", str(e))
        finally:
            connection_minus.close()  # ปิด Connection
    #   and sp.name IN {inv_sql_string}      
    found_invoice_numbers = {item['invoice_number'] for item in inv_minus_one_found}

    # Filtering data_unique_from_excel
    filtered_data = [item for item in data_unique_from_excel if item['invoice_number'] not in found_invoice_numbers]
    
    data_unique_from_excel = filtered_data
        
    return jsonify({
        'data_unique_from_excel': data_unique_from_excel,
        'data_unique_from_query': data_unique_from_query,
        'data_qty_unique': data_qty_unique,
        'data_sku_unique': data_sku_unique,
        'data_not_equal_from_excel':data_not_equal_from_excel,
        'data_not_equal_from_data':data_not_equal_from_data,
        'stamp_invoice_both_have':stamp_invoice_both_have
    })

def generate_excel():
    data_excel = []
    invoice_lines = []
    data = request.get_json()
    dataForExcel1 = data['dataForExcel1']
    dataForExcel2 = data['dataForExcel2']
    dataForExcel1 = ast.literal_eval(dataForExcel1)
    dataForExcel2 = ast.literal_eval(dataForExcel2)

    try:
        connection_invoice = pyodbc.connect(connection_string_invoice)
        cursor_invoice = connection_invoice.cursor()
        emp_code = session['employee_code']
        ############################################################ So query ############################################################
        sql_query = f"""
                        SELECT * FROM invoice_excel
                        WHERE emp_code = ?
                        ORDER BY type_name DESC, invoice_number ASC;
                    """
        cursor_invoice.execute(sql_query, (emp_code,))
        result = cursor_invoice.fetchall()
        data_excel = [{
                    'type_name': str(item.type_name),
                    'invoice_number': str(item.invoice_number),
                    'create_on': str(item.create_on),
                    'barcode': str(item.barcode),
                    'product': str(item.product),
                    'qty': str(item.qty),
                    'weight': str(item.weight),
                    'contact': str(item.contact),
                    'order_type_name': str(item.order_type_name),
                    'internal_reference': str(item.internal_reference),
                    'serial_number': str(item.serial_number),
                    'count': str(item.count),
                    'status_tg': str(item.status_tg)
                } for item in result]
        print("Picking is Success")
    except Exception as e:
        print("Picking Error:", str(e))
    finally:
        connection_invoice.close()  # ปิด Connection
        
    # data_excel
    for data in data_excel:
        invoice_lines.append({
                'type_name': data['type_name'],
                'invoice_number': data['invoice_number'],
                'date': data['create_on'],
                'sku': data['barcode'],
                'material_description': data['product'],
                'qty': data['qty'],
                'weight': data['weight'] if data['weight'] != 'None' or None else '0',
                'supplier_name': data['contact'],
                'doc_type': data['order_type_name'],
                'sloc': data['internal_reference'],
                'serial': data['serial_number'] if str(data['serial_number']) != 'None' or None else '0',
                'status_tg': data['status_tg']
            })


    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # กำหนดรูปแบบ Font และ Border
    font_bold = Font(bold=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # เขียนหัวข้อและข้อมูลลงใน Worksheet
    header = ['ไฟลจาก', 'PO/Invoice', 'Date']
    header2 = ['ไฟลจาก', 'PO/Invoice', 'Date', 'SKU', 'Material description', 'Qty In EWM', 'Weight', 'SupplierName', 'DocType', 'Sloc', 'Serial', 'status_tg']
    ws.append(header)  # เขียนหัวข้อ

    # ตั้งค่าตัวหนาและเส้นกรอบสำหรับแถวหัว
    count_ws = 1
    for cell in ws[1]:
        cell.font = font_bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # เขียนข้อมูลจาก dataForExcel1 และ dataForExcel2
    for row in dataForExcel1:
        count_ws += 1
        ws.append(row)
    for row in dataForExcel2:
        count_ws += 1
        ws.append(row)
        
    # ขยายความยาวของคอลัมน์
    ws.column_dimensions['B'].width = ws.column_dimensions['B'].width * 2
    ws.column_dimensions['C'].width = ws.column_dimensions['C'].width * 2
    ws.column_dimensions['D'].width = ws.column_dimensions['D'].width * 3
    ws.column_dimensions['E'].width = ws.column_dimensions['E'].width * 3
    ws.column_dimensions['F'].width = ws.column_dimensions['F'].width * 1.5
    ws.column_dimensions['H'].width = ws.column_dimensions['H'].width * 3
    ws.column_dimensions['I'].width = ws.column_dimensions['I'].width * 2
    ws.column_dimensions['J'].width = ws.column_dimensions['J'].width * 2
    ws.column_dimensions['K'].width = ws.column_dimensions['K'].width * 3

    
    ws.append(header2)  # เขียนหัวข้อ2
    
    count_ws += 1
    for cell in ws[count_ws]:
        cell.font = font_bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
    for row in invoice_lines:
        count_ws += 1
        row_values = [
            row['type_name'],
            row['invoice_number'],
            row['date'],
            str(row['sku']),
            row['material_description'],
            int(row['qty']) if row['qty'] else None,
            float(row['weight']) if row['weight'] else None,
            row['supplier_name'],
            row['doc_type'],
            row['sloc'],
            row['serial'],
            row['status_tg']
        ]
        ws.append(row_values)
        
        for cell in ws[count_ws]:
            if cell.column_letter in ['F', 'G', 'K']:
                cell.alignment = Alignment(horizontal='right')
                if cell.column_letter in ['F']:
                    cell.number_format = '0'
                if cell.column_letter in ['G']:
                    cell.number_format = '0.0'
    
    # บันทึก Workbook
    wb.save('Data_from_Odoo_(ไม่มีใน_SCG).xlsx')

    # ส่งไฟล์ Excel กลับ
    return send_file('Data_from_Odoo_(ไม่มีใน_SCG).xlsx', as_attachment=True)


def generate_excellines():
    inv_lines = []
    inv_names = []
    inv_odoo_lines = []
    try:
        connection_invoice = pyodbc.connect(connection_string_invoice)
        cursor_invoice = connection_invoice.cursor()
        emp_code = session['employee_code']
        ############################################################ So query ############################################################
        sql_query = f"""
                        SELECT * FROM invoice_excel_line_not_equal
                        WHERE emp_code = ?
                    """
        cursor_invoice.execute(sql_query, (emp_code,))
        result = cursor_invoice.fetchall()
        inv_lines = [{
                    'invoice_number': str(item.invoice_number),
                    'date': str(item.create_on),
                    'sku': str(item.barcode),
                    'material_description': str(item.product),
                    'qty': str(item.qty),
                    'weight': str(item.weight) if str(item.weight) != 'None' or None else '0',
                    'supplier_name': str(item.contact),
                    'doc_type': str(item.order_type_name),
                    'sloc': str(item.internal_reference),
                    'serial': str(item.serial_number) if str(item.serial_number) != 'None' or None else '0',
                    'status_tg': str(item.status_tg)
                } for item in result]
        print("Picking is Success")
    except Exception as e:
        print("Picking Error:", str(e))
    finally:
        connection_invoice.close()  # ปิด Connection
    
    inv_names = [line['invoice_number'] for line in inv_lines]
    invoice_numbers = ','.join([f"'{invoice}'" for invoice in inv_names])
    
    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # กำหนดรูปแบบ Font และ Border
    font_bold = Font(bold=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # เขียนหัวข้อและข้อมูลลงใน Worksheet
    header = ['PO/Invoice', 'File From', 'Date', 'SKU', 'Material description', 'Qty In EWM', 'Weight', 'SupplierName', 'DocType', 'Sloc', 'Serial', 'status_tg']
    ws.append(header)  # เขียนหัวข้อ

    # ตั้งค่าตัวหนาและเส้นกรอบสำหรับแถวหัว
    count_ws = 1
    
    for cell in ws[1]:
        cell.font = font_bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    try:
        connection = pyodbc.connect(connection_string)
        cursor = connection.cursor()
        ############################################################ So query ############################################################
        sql_query = f"""
                        SELECT 
                                sp.name AS invoice_number,
                                TO_CHAR(sp.create_date + interval '7 hours', 'DD.MM.YYYY HH24:MI:SS') AS date,
                                pp.default_code AS sku,
                                pp.product_name AS material_description,
                                ROUND(sm.product_uom_qty) AS qty,
                                pp.weight AS weight,
                                rp.name AS supplier_name,
                                spt.name AS doc_type,
                                sl.internal_reference AS sloc,
                                s_pro_l."name" AS serial,
                                CASE
                                    WHEN sp.state = 'draft' THEN 'Draft'
                                    WHEN sp.state = 'waiting' THEN 'Waiting Another Operation'
                                    WHEN sp.state = 'confirmed' THEN 'Waiting'
                                    WHEN sp.state = 'assigned' THEN 'Ready'
                                    WHEN sp.state = 'done' THEN 'Done'
                                    WHEN sp.state = 'cancel' THEN 'Cancelled'
                                END AS status_tg
                            FROM 
                                stock_picking sp 
                            LEFT JOIN 
                                sale_purchase_types spt ON spt.id = sp.sale_purchase_type_id 
                            LEFT JOIN 
                                stock_location sl ON sl.id = sp.location_id 
                            LEFT JOIN 
                                stock_location sld ON sld.id = sp.location_dest_id 
                            LEFT JOIN 
                                res_partner rp ON rp.id = sp.partner_id 
                            LEFT JOIN 
                                stock_picking_batch spb ON spb.id = sp.batch_id 
                            LEFT JOIN 
                                stock_move sm ON sm.picking_id = sp.id
                            LEFT JOIN 
                                ( 
                                    SELECT 
                                        sml_sub.id,
                                        sml_sub.lot_id,
                                        sml_sub.move_id,
                                        sml_sub.qty_done,
                                        ROW_NUMBER() OVER (PARTITION BY sml_sub.id ORDER BY sml_sub.id) AS rn
                                    FROM
                                        stock_move_line sml_sub
                                ) AS sml ON sml.move_id = sm.id AND rn = 1
                            LEFT JOIN 
                                stock_production_lot s_pro_l ON s_pro_l.id = sml.lot_id                          
                            LEFT JOIN 
                                product_product pp ON pp.id = sm.product_id 
                            WHERE 
                                sp.name IN({invoice_numbers})
                            GROUP BY sp.name,sp.create_date,pp.default_code,pp.product_name,sm.product_uom_qty,pp.weight,rp.name,spt.name,sl.internal_reference,s_pro_l."name",sp.state
                            UNION ALL
                            SELECT 
                                'END_line' AS invoice_number,
                                '' AS date,
                                '' AS sku,
                                '' AS material_description,
                                0 AS qty,
                                0 AS weight,
                                '' AS supplier_name,
                                '' AS doc_type,
                                '' AS sloc,
                                '' AS serial,
                                '' AS status_tg;
                    """
        cursor.execute(sql_query)
        result = cursor.fetchall() 
        inv_odoo_lines = [{
                    'invoice_number': str(item.invoice_number),
                    'date': str(item.date),
                    'sku': str(item.sku),
                    'material_description': str(item.material_description),
                    'qty': str(item.qty),
                    'weight': str(item.weight) if str(item.weight) != 'None' or None else '0',
                    'supplier_name': str(item.supplier_name),
                    'doc_type': str(item.doc_type),
                    'sloc': str(item.sloc),
                    'serial': str(item.serial) if str(item.serial) != 'None' or None else '0',
                    'status_tg': str(item.status_tg)
                } for item in result]
        print("Picking is Success")
    except Exception as e:
        print("Picking Error:", str(e))
    finally:
        connection.close()  # ปิด Connection 
    
    inv_name = ''
    old_inv_name = ''
    is_first = True
    for inv_odoo_line in inv_odoo_lines:
        inv_name = inv_odoo_line['invoice_number']
        if old_inv_name != inv_name:
            if is_first == False:
                row_header_scg = [
                    '',
                    'SCG'
                ]
                ws.append(row_header_scg) 
                for row in inv_lines:
                    if row['invoice_number'] == old_inv_name:
                        row_values_scg = [
                            None,
                            None,
                            row['date'],
                            str(row['sku']),
                            row['material_description'],
                            int(row['qty']) if row['qty'] else None,
                            float(row['weight']) if row['weight'] else None,
                            row['supplier_name'],
                            row['doc_type'],
                            row['sloc'],
                            row['serial'],
                            row['status_tg']
                        ]
                        ws.append(row_values_scg)
                ws.append([])
            row_header_tg = [
                inv_odoo_line['invoice_number'], 
                'TG'
            ]
            if inv_odoo_line['invoice_number'] != 'END_line':
                ws.append(row_header_tg) 
            old_inv_name = inv_name
        row_values = [
            None,
            None,
            inv_odoo_line['date'],
            str(inv_odoo_line['sku']),
            inv_odoo_line['material_description'],
            int(inv_odoo_line['qty']) if inv_odoo_line['qty'] else None,
            float(inv_odoo_line['weight']) if inv_odoo_line['weight'] else None,
            inv_odoo_line['supplier_name'],
            inv_odoo_line['doc_type'],
            inv_odoo_line['sloc'],
            inv_odoo_line['serial'],
            inv_odoo_line['status_tg']
        ]
        if inv_odoo_line['invoice_number'] != 'END_line':
            ws.append(row_values)
        is_first = False

    # ขยายความยาวของคอลัมน์
    ws.column_dimensions['A'].width = ws.column_dimensions['A'].width * 3
    ws.column_dimensions['C'].width = ws.column_dimensions['C'].width * 2
    ws.column_dimensions['D'].width = ws.column_dimensions['D'].width * 2
    ws.column_dimensions['E'].width = ws.column_dimensions['E'].width * 3
    ws.column_dimensions['F'].width = ws.column_dimensions['F'].width * 1
    ws.column_dimensions['G'].width = ws.column_dimensions['G'].width * 1
    ws.column_dimensions['H'].width = ws.column_dimensions['H'].width * 3


    # บันทึก Workbook
    wb.save('Invoice ที่ตรงกันแต่ Line ไม่ตรงกัน.xlsx')

    # ส่งไฟล์ Excel กลับ
    return send_file('Invoice ที่ตรงกันแต่ Line ไม่ตรงกัน.xlsx', as_attachment=True)
